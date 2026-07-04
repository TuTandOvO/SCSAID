#!/usr/bin/env python3
"""Build the scSAID dataset-to-publication registry.

The source workbook provides the canonical SAID/study/sample mapping. PubMed XML
is supplied explicitly so generation is deterministic and does not require
network access.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import unicodedata
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = ROOT.parent / "src/main/webapp/WEB-INF/AllData.xlsx"
OVERRIDES = ROOT / "study_publication_overrides.json"


def text(node: ET.Element | None) -> str:
    return "" if node is None else "".join(node.itertext()).strip()


def first(*values: str | None) -> str:
    return next((value.strip() for value in values if value and value.strip()), "")


def normalize_title(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return " ".join(re.findall(r"[a-z0-9]+", value.lower()))


def title_similarity(left: str, right: str) -> float:
    return SequenceMatcher(None, normalize_title(left), normalize_title(right)).ratio()


def pub_date(article: ET.Element) -> str:
    node = article.find(".//JournalIssue/PubDate")
    if node is None:
        return ""
    medline = node.findtext("MedlineDate")
    if medline:
        match = re.search(r"\d{4}", medline)
        return match.group(0) if match else medline
    year = node.findtext("Year") or ""
    month = node.findtext("Month") or ""
    day = node.findtext("Day") or ""
    return "-".join(part for part in (year, month, day) if part)


def parse_pubmed(path: Path) -> dict[str, dict]:
    records: dict[str, dict] = {}
    for entry in ET.parse(path).getroot().findall("PubmedArticle"):
        pmid = first(entry.findtext(".//MedlineCitation/PMID"))
        if not pmid:
            continue
        article = entry.find(".//Article")
        if article is None:
            continue
        authors = []
        for author in article.findall(".//AuthorList/Author"):
            collective = first(author.findtext("CollectiveName"))
            if collective:
                authors.append(collective)
                continue
            display = " ".join(
                part for part in (
                    first(author.findtext("ForeName"), author.findtext("Initials")),
                    first(author.findtext("LastName")),
                ) if part
            )
            if display:
                authors.append(display)
        article_ids = {
            item.attrib.get("IdType", "").lower(): first(item.text)
            for item in entry.findall(".//PubmedData/ArticleIdList/ArticleId")
        }
        journal = first(article.findtext(".//Journal/Title"))
        journal_abbrev = first(entry.findtext(".//MedlineJournalInfo/MedlineTA"))
        publication_types = [text(item) for item in article.findall(".//PublicationTypeList/PublicationType")]
        record = {
            "paper_id": f"PMID_{pmid}",
            "pmid": pmid,
            "pmcid": article_ids.get("pmc", ""),
            "doi": article_ids.get("doi", ""),
            "title": text(article.find("ArticleTitle")),
            "authors": authors,
            "journal": journal,
            "journal_abbreviation": journal_abbrev,
            "publication_date": pub_date(article),
            "year": (pub_date(article)[:4] if pub_date(article) else ""),
            "volume": first(article.findtext(".//JournalIssue/Volume")),
            "issue": first(article.findtext(".//JournalIssue/Issue")),
            "pages": first(article.findtext("Pagination/MedlinePgn"), article.findtext("ELocationID")),
            "publication_types": publication_types,
            "language": first(article.findtext("Language")),
            "urls": {
                "pubmed": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/",
                "doi": f"https://doi.org/{article_ids['doi']}" if article_ids.get("doi") else "",
                "pmc": f"https://pmc.ncbi.nlm.nih.gov/articles/{article_ids['pmc']}/" if article_ids.get("pmc") else "",
            },
        }
        records[pmid] = record
    return records


def load_datasets(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(value) for value in next(sheet.iter_rows(values_only=True))]
    positions = {name: index for index, name in enumerate(headers)}
    required = {"SAID", "GSE", "GSM", "Title", "PubMed"}
    missing = required - positions.keys()
    if missing:
        raise ValueError(f"Workbook lacks columns: {', '.join(sorted(missing))}")
    datasets = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        datasets.append({
            "said": first(str(row[positions["SAID"]] or "")),
            "study_accession": first(str(row[positions["GSE"]] or "")),
            "sample_accession": first(str(row[positions["GSM"]] or "")),
            "study_title": first(str(row[positions["Title"]] or "")),
            "workbook_pmids": re.findall(r"\d+", str(row[positions["PubMed"]] or "")),
        })
    return datasets


def citation(record: dict) -> str:
    authors = record["authors"]
    author_text = ", ".join(authors[:6])
    if len(authors) > 6:
        author_text += ", et al."
    journal = record["journal_abbreviation"] or record["journal"]
    chunks = [f"{author_text}." if author_text else "", f"{record['title']}", f"{journal}." if journal else ""]
    if record["year"]:
        chunks.append(record["year"] + ".")
    if record["doi"]:
        chunks.append(f"doi:{record['doi']}.")
    chunks.append(f"PMID:{record['pmid']}.")
    return " ".join(chunk for chunk in chunks if chunk).replace("..", ".")


def write_tsv(path: Path, rows: list[dict], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
            delimiter="\t",
            extrasaction="ignore",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--pubmed-xml", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=ROOT)
    parser.add_argument("--verified-on", default=date.today().isoformat())
    args = parser.parse_args()

    papers = parse_pubmed(args.pubmed_xml)
    datasets = load_datasets(args.workbook)
    overrides = json.loads(OVERRIDES.read_text(encoding="utf-8"))
    output = args.output.resolve()
    paper_dir = output / "papers"
    if paper_dir.exists():
        shutil.rmtree(paper_dir)
    paper_dir.mkdir(parents=True)

    studies: dict[str, list[dict]] = defaultdict(list)
    for dataset in datasets:
        studies[dataset["study_accession"]].append(dataset)

    study_links: dict[str, dict[str, dict]] = defaultdict(dict)
    for accession, members in studies.items():
        study_title = next((item["study_title"] for item in members if item["study_title"]), "")
        workbook_pmids = {pmid for item in members for pmid in item["workbook_pmids"]}
        for pmid in workbook_pmids:
            record = papers.get(pmid)
            if not record:
                raise ValueError(f"PMID {pmid} from {accession} is absent from PubMed XML")
            similarity = title_similarity(study_title, record["title"]) if study_title else 0
            study_links[accession][pmid] = {
                "relation": "primary_dataset_publication" if similarity >= 0.72 else "repository_linked_publication",
                "evidence": "PubMed identifier recorded in AllData.xlsx",
                "provenance": "AllData.xlsx",
                "title_similarity": f"{similarity:.3f}",
            }
        for item in overrides.get(accession, []):
            pmid = item["pmid"]
            if pmid not in papers:
                raise ValueError(f"Override PMID {pmid} for {accession} is absent from PubMed XML")
            study_links[accession][pmid] = {
                "relation": item["relation"],
                "evidence": item["evidence"],
                "provenance": "study_publication_overrides.json",
                "title_similarity": "",
            }

    study_rows = []
    dataset_rows = []
    paper_datasets: dict[str, list[dict]] = defaultdict(list)
    for accession in sorted(studies):
        members = studies[accession]
        for pmid in sorted(study_links[accession], key=int):
            link = study_links[accession][pmid]
            record = papers[pmid]
            base = {
                "study_accession": accession,
                "paper_id": record["paper_id"],
                "pmid": pmid,
                "doi": record["doi"],
                **link,
                "verified_on": args.verified_on,
            }
            study_rows.append({**base, "dataset_count": len(members)})
            for member in members:
                row = {
                    "said": member["said"],
                    "study_accession": accession,
                    "sample_accession": member["sample_accession"],
                    "study_title": member["study_title"],
                    **{key: value for key, value in base.items() if key != "study_accession"},
                }
                dataset_rows.append(row)
                paper_datasets[pmid].append(row)

    used_pmids = {row["pmid"] for row in study_rows}
    paper_rows = []
    for pmid in sorted(used_pmids, key=int):
        record = papers[pmid]
        links = paper_datasets[pmid]
        record["citation"] = citation(record)
        record["dataset_links"] = [
            {
                "said": row["said"],
                "study_accession": row["study_accession"],
                "sample_accession": row["sample_accession"],
                "relation": row["relation"],
            }
            for row in links
        ]
        destination = paper_dir / record["paper_id"]
        destination.mkdir()
        (destination / "metadata.json").write_text(json.dumps(record, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        source_links = [f"- [PubMed]({record['urls']['pubmed']})"]
        if record["urls"]["doi"]:
            source_links.append(f"- [Publisher / DOI]({record['urls']['doi']})")
        if record["urls"]["pmc"]:
            source_links.append(f"- [Open full text in PubMed Central]({record['urls']['pmc']})")
        related = sorted({(row["study_accession"], row["relation"]) for row in links})
        readme = (
            f"# {record['title']}\n\n{record['citation']}\n\n"
            "## Sources\n\n" + "\n".join(source_links) + "\n\n"
            "## scSAID dataset relations\n\n" +
            "\n".join(f"- `{accession}` — `{relation}`" for accession, relation in related) + "\n"
        )
        (destination / "README.md").write_text(readme, encoding="utf-8")
        paper_rows.append({
            "paper_id": record["paper_id"], "pmid": pmid, "pmcid": record["pmcid"],
            "doi": record["doi"], "title": record["title"], "authors": "; ".join(record["authors"]),
            "journal": record["journal"], "year": record["year"],
            "publication_types": "; ".join(record["publication_types"]),
            "study_count": len({row["study_accession"] for row in links}),
            "dataset_count": len(links), "pubmed_url": record["urls"]["pubmed"],
            "full_text_url": record["urls"]["pmc"] or record["urls"]["doi"],
        })

    unresolved = []
    for accession, members in sorted(studies.items()):
        if study_links[accession]:
            continue
        unresolved.append({
            "study_accession": accession,
            "study_title": next((item["study_title"] for item in members if item["study_title"]), ""),
            "dataset_count": len(members),
            "status": "no_primary_publication_verified",
            "checked_on": args.verified_on,
        })

    write_tsv(output / "papers.tsv", paper_rows, list(paper_rows[0]))
    write_tsv(output / "study_paper_index.tsv", study_rows, list(study_rows[0]))
    write_tsv(output / "dataset_paper_index.tsv", dataset_rows, list(dataset_rows[0]))
    write_tsv(output / "unresolved_studies.tsv", unresolved, ["study_accession", "study_title", "dataset_count", "status", "checked_on"])
    registry = {
        "schema_version": 1,
        "generated_on": args.verified_on,
        "counts": {
            "datasets": len(datasets), "studies": len(studies), "papers": len(paper_rows),
            "dataset_paper_links": len(dataset_rows), "unresolved_studies": len(unresolved),
        },
        "papers": paper_rows,
        "dataset_paper_links": dataset_rows,
        "unresolved_studies": unresolved,
    }
    (output / "registry.json").write_text(json.dumps(registry, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(registry["counts"], sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
